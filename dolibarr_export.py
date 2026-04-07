"""
Dolibarr Facturación Excel — AutomaWorks
Genera un Excel con la facturación mensual desde la API de Dolibarr.
"""

import flet as ft
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import urllib.request
import io
import os
import tempfile

# ── Paleta AutomaWorks Light ──────────────────────────────────────────────────
BG        = "#f5f7fb"
BG2       = "#ffffff"
TEXT      = "#0f172a"
TEXT2     = "#4b5c78"
ACCENT    = "#0891b2"
ACCENT_DIM= "#e0f2f7"
BORDER    = "#e2e8f0"
RADIUS    = 10
BG_HEADER = "#0f1d2e"

LOGO_URL  = "automaworks_logo.png"
WEB_URL   = "https://automaworks.es"
EMAIL     = "iflorido@gmail.com"

# ── Helpers ───────────────────────────────────────────────────────────────────

def parse_date(s: str) -> datetime | None:
    """Acepta DD-MM-YYYY"""
    try:
        return datetime.strptime(s.strip(), "%d-%m-%Y")
    except ValueError:
        return None

def fetch_invoices(base_url: str, api_key: str, date_from: datetime, date_to: datetime):
    """Obtiene facturas de Dolibarr en el rango indicado."""
    url = base_url.rstrip("/") + "/api/index.php/invoices"
    headers = {"DOLAPIKEY": api_key, "Accept": "application/json"}

    ts_from = int(date_from.timestamp())
    ts_to   = int(date_to.replace(hour=23, minute=59, second=59).timestamp())

    params = {
        "limit": 500,
        "page":  0,
        "sqlfilters": f"(t.date_lim_reglement:>=:{ts_from}) and (t.date_lim_reglement:<=:{ts_to})",
    }

    # Intentamos filtrar por fecha de factura directamente
    params_date = {
        "limit": 500,
        "page":  0,
        "sqlfilters": f"(t.datef:>=:{ts_from}) and (t.datef:<=:{ts_to})",
    }

    try:
        r = requests.get(url, headers=headers, params=params_date, timeout=15)
        r.raise_for_status()
        data = r.json()
        if isinstance(data, list):
            return data
        return []
    except Exception as e:
        raise RuntimeError(f"Error al conectar con Dolibarr: {e}")


def get_third_party(base_url: str, api_key: str, thirdparty_id: int) -> dict:
    """Obtiene datos del tercero (empresa/cliente)."""
    url = base_url.rstrip("/") + f"/api/index.php/thirdparties/{thirdparty_id}"
    headers = {"DOLAPIKEY": api_key, "Accept": "application/json"}
    try:
        r = requests.get(url, headers=headers, timeout=10)
        r.raise_for_status()
        return r.json()
    except Exception:
        return {}


def build_excel(invoices: list, base_url: str, api_key: str,
                date_from: datetime, date_to: datetime) -> bytes:
    """Construye el Excel y devuelve los bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturación"

    # Paleta
    accent_hex  = "0891B2"
    header_fill = PatternFill("solid", fgColor=accent_hex)
    alt_fill    = PatternFill("solid", fgColor="F0F9FF")
    white_fill  = PatternFill("solid", fgColor="FFFFFF")
    thin        = Side(style="thin", color="CBD5E1")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Título
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = (
        f"Facturación  {date_from.strftime('%d/%m/%Y')} → {date_to.strftime('%d/%m/%Y')}"
    )
    title_cell.font      = Font(name="Calibri", bold=True, size=14, color="0F172A")
    title_cell.fill      = PatternFill("solid", fgColor="E0F2F7")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Generado por AutomaWorks — {WEB_URL}"
    sub_cell.font  = Font(name="Calibri", size=9, color="4B5C78", italic=True)
    sub_cell.alignment = Alignment(horizontal="center")

    # Cabeceras
    headers = [
        "Fecha Factura", "Nº Factura", "Empresa",
        "CIF", "Dirección", "Base (€)", "IVA (€)", "Total (€)"
    ]
    # Ajustamos merge al rango real (8 columnas)
    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")

    row3 = 3
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row3, column=col, value=h)
        cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[row3].height = 22

    # Datos
    thirdparty_cache: dict[int, dict] = {}
    row_num = row3 + 1
    total_base = total_iva = total_total = 0.0

    for inv in invoices:
        tp_id = inv.get("socid") or inv.get("thirdparty_id") or 0
        if tp_id and tp_id not in thirdparty_cache:
            thirdparty_cache[tp_id] = get_third_party(base_url, api_key, int(tp_id))
        tp = thirdparty_cache.get(tp_id, {})

        # Fecha
        ts = inv.get("date") or inv.get("datef") or 0
        try:
            fecha = datetime.fromtimestamp(int(ts)).strftime("%d/%m/%Y") if ts else ""
        except Exception:
            fecha = ""

        numero   = inv.get("ref", "")
        empresa  = tp.get("name", inv.get("socnom", ""))
        cif      = tp.get("idprof2", tp.get("siren", ""))
        ciudad   = tp.get("town", "")
        cp       = tp.get("zip", "")
        pais     = tp.get("country", {}).get("label", "") if isinstance(tp.get("country"), dict) else ""
        direccion_parts = [tp.get("address", ""), f"{cp} {ciudad}".strip(), pais]
        direccion = ", ".join(p for p in direccion_parts if p)

        base  = float(inv.get("total_ht",  0) or 0)
        iva   = float(inv.get("total_tva", 0) or 0)
        total = float(inv.get("total_ttc", 0) or 0)

        total_base  += base
        total_iva   += iva
        total_total += total

        fill = alt_fill if (row_num % 2 == 0) else white_fill
        row_data = [fecha, numero, empresa, cif, direccion, base, iva, total]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font      = Font(name="Calibri", size=9, color="0F172A")
            cell.fill      = fill
            cell.border    = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 5))
            if col in (6, 7, 8):
                cell.number_format = '#,##0.00 €'
                cell.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[row_num].height = 16
        row_num += 1

    # Fila totales
    total_row = row_num
    ws.cell(total_row, 1, "TOTALES").font = Font(bold=True, name="Calibri", size=10, color="FFFFFF")
    ws.cell(total_row, 1).fill   = header_fill
    ws.cell(total_row, 1).border = border
    ws.cell(total_row, 1).alignment = Alignment(horizontal="right")

    for col in range(2, 6):
        c = ws.cell(total_row, col, "")
        c.fill   = header_fill
        c.border = border

    for col, val in zip([6, 7, 8], [total_base, total_iva, total_total]):
        c = ws.cell(total_row, col, val)
        c.font          = Font(bold=True, name="Calibri", size=10, color="FFFFFF")
        c.fill          = header_fill
        c.border        = border
        c.number_format = '#,##0.00 €'
        c.alignment     = Alignment(horizontal="right", vertical="center")

    ws.row_dimensions[total_row].height = 20

    # Anchos de columna
    col_widths = [13, 14, 28, 14, 36, 13, 12, 13]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── UI ────────────────────────────────────────────────────────────────────────

def main(page: ft.Page):
    page.title       = "Dolibarr · Facturación Excel"
    page.window.width  = 680
    page.window.height = 560
    page.window.resizable = False
    page.bgcolor     = BG
    page.padding     = 0
    page.fonts       = {}
    page.theme_mode  = ft.ThemeMode.LIGHT

    # ── Estado ────────────────────────────────────────────────────────────────
    status_text = ft.Text("", size=12, color=TEXT2, text_align=ft.TextAlign.CENTER)
    progress    = ft.ProgressBar(visible=False, color=ACCENT, bgcolor=ACCENT_DIM)

    def show_status(msg: str, color: str = TEXT2, loading: bool = False):
        status_text.value   = msg
        status_text.color   = color
        progress.visible    = loading
        page.update()

    # ── Logo ──────────────────────────────────────────────────────────────────
    logo = ft.Image(
        src=LOGO_URL,
        height=38,
        fit="contain",
        error_content=ft.Text(
            "AutomaWorks", size=18, weight=ft.FontWeight.BOLD, color=ACCENT
        ),
    )

    header = ft.Container(
        content=ft.Row([logo], alignment=ft.MainAxisAlignment.START),
        bgcolor=BG_HEADER,
        padding=ft.padding.symmetric(vertical=14, horizontal=20),
        border=ft.border.only(bottom=ft.BorderSide(1, BORDER)),
    )

    # ── Campos ────────────────────────────────────────────────────────────────
    def field(label: str, hint: str, password: bool = False, value: str = "") -> ft.TextField:
        return ft.TextField(
            label=label,
            hint_text=hint,
            password=password,
            can_reveal_password=password,
            value=value,
            bgcolor=BG2,
            border_color=BORDER,
            focused_border_color=ACCENT,
            label_style=ft.TextStyle(color=TEXT2, size=12),
            text_style=ft.TextStyle(color=TEXT, size=13),
            hint_style=ft.TextStyle(color=TEXT2, size=12),
            border_radius=RADIUS,
            content_padding=ft.padding.symmetric(horizontal=12, vertical=10),
            dense=True,
        )

    tf_url   = field("URL de Dolibarr", "https://midominio.com/dolibarr")
    tf_key   = field("API Key", "••••••••••••••••", password=True)
    tf_from  = field("Fecha inicio", "DD-MM-YYYY")
    tf_to    = field("Fecha final",  "DD-MM-YYYY")

    # ── Botón generar ─────────────────────────────────────────────────────────
    def on_generate(e):
        url_val  = tf_url.value.strip()
        key_val  = tf_key.value.strip()
        from_val = tf_from.value.strip()
        to_val   = tf_to.value.strip()

        # Validaciones
        if not url_val or not key_val or not from_val or not to_val:
            show_status("⚠️  Completa todos los campos.", "#e11d48")
            return

        d_from = parse_date(from_val)
        d_to   = parse_date(to_val)
        if not d_from:
            show_status("⚠️  Fecha inicio inválida (DD-MM-YYYY).", "#e11d48")
            return
        if not d_to:
            show_status("⚠️  Fecha final inválida (DD-MM-YYYY).", "#e11d48")
            return
        if d_to < d_from:
            show_status("⚠️  La fecha final debe ser posterior a la inicial.", "#e11d48")
            return

        btn_generate.disabled = True
        show_status("Consultando facturas en Dolibarr…", TEXT2, loading=True)

        try:
            invoices = fetch_invoices(url_val, key_val, d_from, d_to)
            if not invoices:
                show_status("ℹ️  No se encontraron facturas en ese rango.", TEXT2)
                btn_generate.disabled = False
                page.update()
                return

            show_status(f"Generando Excel con {len(invoices)} facturas…", TEXT2, loading=True)
            excel_bytes = build_excel(invoices, url_val, key_val, d_from, d_to)

            # Guardar en Escritorio del usuario
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            os.makedirs(desktop, exist_ok=True)
            fname = f"facturacion_{d_from.strftime('%Y%m%d')}_{d_to.strftime('%Y%m%d')}.xlsx"
            fpath = os.path.join(desktop, fname)
            with open(fpath, "wb") as f:
                f.write(excel_bytes)

            show_status(f"✅  Excel guardado en Escritorio:\n{fname}", "#059669")
        except Exception as ex:
            show_status(f"❌  {ex}", "#e11d48")
        finally:
            btn_generate.disabled = False
            progress.visible = False
            page.update()

    btn_generate = ft.ElevatedButton(
        content=ft.Row(
            [
                ft.Icon(ft.Icons.TABLE_CHART_OUTLINED, color="#ffffff", size=18),
                ft.Text("Generar Excel", color="#ffffff", size=13,
                        weight=ft.FontWeight.W_600),
            ],
            alignment=ft.MainAxisAlignment.CENTER,
            spacing=8,
            tight=True,
        ),
        on_click=on_generate,
        bgcolor=ACCENT,
        style=ft.ButtonStyle(
            shape=ft.RoundedRectangleBorder(radius=RADIUS),
            padding=ft.padding.symmetric(vertical=12, horizontal=20),
        ),
        width=220,
    )

    # ── Cuerpo ────────────────────────────────────────────────────────────────
    body = ft.Container(
        content=ft.Column(
            controls=[
                ft.Text("Exportar Facturación", size=16,
                        weight=ft.FontWeight.BOLD, color=TEXT),
                ft.Text("Conecta con tu Dolibarr y descarga el Excel.",
                        size=12, color=TEXT2),
                ft.Divider(height=4, color="transparent"),
                tf_url,
                tf_key,
                ft.Row([tf_from, tf_to], spacing=10),
                ft.Divider(height=2, color="transparent"),
                ft.Row([btn_generate], alignment=ft.MainAxisAlignment.CENTER),
                progress,
                status_text,
            ],
            spacing=10,
            scroll=ft.ScrollMode.AUTO,
        ),
        padding=ft.padding.symmetric(horizontal=24, vertical=20),
        expand=True,
        bgcolor=BG,
    )

    # ── Footer ────────────────────────────────────────────────────────────────
    footer = ft.Container(
        content=ft.Row(
            [
                ft.Text("© AutomaWorks · ", size=10, color=TEXT2),
                ft.TextButton(
                    content=ft.Text(WEB_URL, size=10, color=ACCENT),
                    url=WEB_URL,
                    style=ft.ButtonStyle(padding=ft.padding.all(0)),
                ),
                ft.Text("  ·  ", size=10, color=TEXT2),
                ft.TextButton(
                    content=ft.Text(EMAIL, size=10, color=ACCENT),
                    url=f"mailto:{EMAIL}",
                    style=ft.ButtonStyle(padding=ft.padding.all(0)),
                ),
            ],
            alignment=ft.MainAxisAlignment.CENTER,
            spacing=0,
        ),
        bgcolor=BG2,
        padding=ft.padding.symmetric(vertical=8, horizontal=16),
        border=ft.border.only(top=ft.BorderSide(1, BORDER)),
    )

    # ── Layout principal ──────────────────────────────────────────────────────
    page.add(
        ft.Column(
            controls=[header, body, footer],
            spacing=0,
            expand=True,
        )
    )

if __name__ == "__main__":
    assets = os.path.dirname(os.path.abspath(__file__))
    ft.app(target=main, assets_dir=assets)